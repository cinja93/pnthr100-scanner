#!/usr/bin/env python3
"""Parse a NAV (or PNTHR-rebranded) Fund Accounting workbook's "Account Statement" tab
into the engine `inputs` contract + NAV's own stored totals (for the reconciliation gate).

Rebranding only recolors cells; it never changes values, so this parses the SAME numbers
whether the workbook is the NAV original or the PNTHR-rebranded copy — which keeps the
two investor PDFs reproducible from whatever is stored.

Usage:  navAccountStatementParser.py <workbook.xlsx> <period YYYY-MM>   # prints JSON to stdout
"""
import openpyxl, json, sys

# Income-statement line item label -> engine inputs.lineItems field.
FIELD = {
 'Realized P&L - Short Term':'realizedPL','Change In Unrealized P&L':'unrealizedPL',
 'Commission Expenses':'commission','Other Trading Cost':'otherTradingCost',
 'Broker Interest Income':'brokerInterestIncome','Dividend Income - US Stock':'divIncomeUS',
 'Dividend Income - Foreign Stock':'divIncomeForeign','Broker Interest Expense':'brokerInterestExpense',
 'Dividend Expense - US Stock':'divExpenseUS','Dividend Expense - Foreign Stock':'divExpenseForeign',
 'Administration Expenses':'admin','Legal Expenses':'legal','Professional Expenses':'professional',
 'Operating Expenses':'operating','Organization Cost':'orgCost','Reimbursement to/from Affiliates':'reimbursement',
}
# NAV's own computed totals -> reconciliation-gate keys (engine must reproduce these to the penny).
TOTALS = {'Total Income (Loss):':'totalIncome','Total Expenses:':'totalExpenses',
 'Net Income (Loss) :':'netIncome','Ending Balance:':'ending','NET ROR:':'ror'}

def vals(ws, r):
    return [ws.cell(r,c).value for c in (2,3,4,5)]   # B,C,D,E = PTD,MTD,QTD,YTD

def parse(period, fp):
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb['Account Statement']
    def s(coord):
        v = ws[coord].value
        return (str(v).strip() if v is not None else '')
    # Header (fixed rows 1-5 — never shifted by omitted income lines).
    header = {
      'fundName': s('A2'), 'statementTitle': s('A3'),
      'periodEnded': s('A4').replace('For the Period Ended ','').strip(),
      'currency': s('A5').replace('Reporting Currency :','').strip(),
      'startOfPeriod': s('C4').replace('Start Of Period :','').strip(),
      'endOfPeriod': s('C5').replace('End Of Period','').replace(':','').strip(),
    }
    lineItems = {f:[None,None,None,None] for f in FIELD.values()}
    beginning=[None]*4; additions=[None]*4; redemptions=[None]*4
    navStored={}; signatory=[]
    for r in range(6, ws.max_row+1):
        a = ws.cell(r,1).value
        if a is None: continue
        label = str(a).strip()
        if label in FIELD: lineItems[FIELD[label]] = vals(ws,r)
        elif label in TOTALS: navStored[TOTALS[label]] = vals(ws,r)
        elif label == 'Beginning Balance': beginning = vals(ws,r)
        elif label == 'Additions': additions = vals(ws,r)
        elif label == 'Redemptions': redemptions = vals(ws,r)
        elif 'General Partner of' in label or label.startswith('For PNTHR') or label.startswith('For NAV'):
            signatory = [x.strip() for x in str(a).split('\n') if x.strip()]
    return {'period':period,'header':header,'lineItems':lineItems,
            'beginning':beginning,'additions':additions,'redemptions':redemptions,
            'signatory':signatory,'navStored':navStored}

if __name__ == '__main__':
    print(json.dumps(parse(sys.argv[2], sys.argv[1])))
