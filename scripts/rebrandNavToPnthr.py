#!/usr/bin/env python3
"""Re-brand a NAV-produced .xlsx into PNTHR branding (black + PNTHR yellow #FCF000).

- Removes NAV logo/band images.
- Recolors NAV navy (FF1F4E78) fills -> black, with PNTHR-yellow band lettering.
- Recolors navy fonts on white -> black ink.
- Sets every worksheet tab color to black.
- Surgically swaps NAV-the-company text only (leaves 'Net Asset Value' etc. intact);
  replaces the disclaimer with PNTHR wording.
- Adds the PNTHR (black-background) logo to each sheet header.
Usage: rebrandNavToPnthr.py <in.xlsx> <out.xlsx> <funds|notebook|caproll>
"""
import openpyxl, os, sys
from openpyxl.styles import PatternFill, Font, Color, Border, Side, Alignment
from openpyxl.drawing.image import Image as XLImage

def _is_black(cell):
    try:
        return cell.fill and cell.fill.fill_type=='solid' and cell.fill.fgColor and cell.fill.fgColor.rgb=='FF000000'
    except Exception:
        return False

NAVY='FF1F4E78'; BLACK='FF000000'; YELLOW='FFFCF000'   # PNTHR yellow sampled from the logo
LOGO='/Users/cindyeagar/pnthr100-scanner/client/public/pnthr-logo-black-bg.png'  # head + wordmark on black
DISCLAIMER=('Disclaimer:\n\nThese statements are prepared and provided by PNTHR Funds, LLC, the General Partner of the Fund, '
 'solely for informational purposes. The figures are unaudited estimates as of the reporting date and are subject to change; '
 'they are superseded by the Fund’s audited annual financial statements. This information is confidential, intended only for the '
 'named investor, and may not be redistributed. It does not constitute an offer, solicitation, or tax, legal, or investment advice. '
 'Past performance is not necessarily indicative of future results.')
REPL=[('NAV Fund Services','PNTHR Funds, LLC'),('NAV Fund Service','PNTHR Funds, LLC'),('NAV Consulting, Inc.','PNTHR Funds, LLC'),
 ('NAV Consulting','PNTHR Funds, LLC'),('NAV Fund Administration Group','PNTHR Funds, LLC'),(' (NAV)',''),('used by NAV to','used to'),
 ('navfundservices.com','pnthrfunds.com'),('navconsulting.net','pnthrfunds.com')]
def fix_text(v):
    if not isinstance(v,str) or 'NAV' not in v: return v
    if v.strip().startswith('Disclaimer'): return DISCLAIMER
    for a,b in REPL: v=v.replace(a,b)
    return v.replace('NAV','PNTHR')
def set_font(cell, rgb, bold=None):
    f=cell.font
    cell.font=Font(name=f.name, size=f.size, bold=(f.bold if bold is None else bold), italic=f.italic, color=Color(rgb=rgb))
def rebrand(infile, outfile, kind):
    wb=openpyxl.load_workbook(infile)
    for ws in wb.worksheets:
        ws.sheet_properties.tabColor='000000'          # black sheet tab
        ws._images=[]                                   # drop NAV logo/band images
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value,str) and 'NAV' in c.value: c.value=fix_text(c.value)
                try:
                    if c.fill and c.fill.fill_type=='solid' and c.fill.fgColor and c.fill.fgColor.rgb==NAVY:
                        c.fill=PatternFill(start_color=BLACK,end_color=BLACK,fill_type='solid'); set_font(c, YELLOW)
                    elif c.font and c.font.color and c.font.color.rgb==NAVY:
                        set_font(c, BLACK)
                except Exception: pass
        # Notebook/CapRoll lost their image band -> build a black band (rows 1-2) w/ yellow fund name
        if kind in ('notebook','caproll'):
            maxc=min(max(ws.max_column,1),40)
            for r in (1,2):
                for cc in range(1,maxc+1):
                    ws.cell(r,cc).fill=PatternFill(start_color=BLACK,end_color=BLACK,fill_type='solid')
            if isinstance(ws.cell(2,1).value,str): set_font(ws.cell(2,1), YELLOW, bold=True)
        # Give row 1 height for the logo (so it sits ABOVE the fund-name text, no overlap)
        ws.row_dimensions[1].height=40
        # Frame the black band with a PNTHR-yellow outline (top/left/right/bottom)
        side=Side(style='medium', color='FCF000')
        last_row=0
        for r in range(1,9):
            if _is_black(ws.cell(r,1)): last_row=r
            else: break
        last_col=0
        for c in range(1, min(max(ws.max_column,1),60)+1):
            if _is_black(ws.cell(1,c)): last_col=c
            else: break
        for r in range(1,last_row+1):
            for c in range(1,last_col+1):
                sides={}
                if r==1: sides['top']=side
                if r==last_row: sides['bottom']=side
                if c==1: sides['left']=side
                if c==last_col: sides['right']=side
                if sides: ws.cell(r,c).border=Border(**sides)
        # Indent any text in row 1 so it clears the logo (e.g. ReportLinks fund name in row 1)
        for c in range(1, max(last_col,1)+1):
            cell=ws.cell(1,c)
            if isinstance(cell.value,str) and cell.value.strip():
                a=cell.alignment
                cell.alignment=Alignment(indent=13, horizontal=(a.horizontal or 'left'), vertical=a.vertical, wrap_text=a.wrap_text)
        if os.path.exists(LOGO):
            try:
                im=XLImage(LOGO); im.height=36; im.width=int(36*800/325); ws.add_image(im,'A1')
            except Exception: pass
    wb.save(outfile)
if __name__=='__main__':
    rebrand(sys.argv[1], sys.argv[2], sys.argv[3])
